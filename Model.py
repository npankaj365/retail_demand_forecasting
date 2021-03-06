import sys
import math
import numpy as np
import pandas as pd
from dateutil import parser as dtparser
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from statsmodels.tsa.arima_model import ARIMA
from sklearn import model_selection
from sklearn.metrics import mean_squared_error, mean_absolute_error 

np.warnings.filterwarnings('ignore')

class Model():
    name = ''

    def __init__(self, file, class_no):
        print("Started {} Model".format(self.name))
        self.data = self.load_data(file)
        print("Completed Loading")
        self.export_excel(self.data, class_no)
        print("Exporting to Excel for Class {}".format(class_no))
        self.select_data(class_no)
        print("Completed Selecting")
        self.intrapolate_data()
        print("Completed Filling")
        self.split_data()
        print("Completed Splitting")
        self.core_model()
        print("Completed {} Model".format(self.name))

    def load_data(self, file):
        csv_data = pd.read_csv(file,
                       index_col='Day',
                       parse_dates=['Day', 'Fiscal Week'],
                       usecols = ['Fiscal Season','SeasonDesc','Fiscal Year','Fiscal Week','Day','Dayofweek','Class','ClassDesc','Location','Locdesc','SalesU','SalesD'],
                       date_parser=lambda s: dtparser.parse(s).date())
        return csv_data

    def export_excel(self, csv_data, class_no):
        df = csv_data.loc[csv_data['Class'] == int(class_no)]
        #Saving to Excel
        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
            
        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        wb.save("result.xlsx")

    def select_data(self, class_no):
        self.sales_data = self.data.query('Class=={}'.format(class_no)).SalesD.astype('float')

    def intrapolate_data(self):
        # Fill in missing values (some dates are missing in the index)
        start_date = self.sales_data.index[0].date()
        end_date = self.sales_data.index[-1].date()
        date_range = pd.date_range(start_date, end_date)
        self.X = self.sales_data.reindex(date_range, fill_value=0).values

    def split_data(self):
        # self.X_train, self.X_test = model_selection.train_test_split(self.X, test_size = 0.2)
        size = int(len(self.X) * 0.66)
        self.X_train, self.X_test = self.X[0:size], self.X[size:len(self.X)]
        print("Training Set size: {} and Test Set size: {}".format(self.X_train.size, self.X_test.size))
        self.X_train = list(self.X_train)

    def core_model(self):
        pass

    def get_mse(self, test_data, predicted_data):
        return mean_squared_error(test_data, predicted_data)

    def get_mae(self, test_data, predicted_data):
        return mean_absolute_error(test_data, predicted_data)

    def get_mape(self, test_data, predicted_data): 
        test_data, predicted_data = np.array(test_data), np.array(predicted_data)
        return np.mean(np.abs((test_data - predicted_data) / test_data)) * 100

class ARIMA_Model(Model):

    name = 'ARIMA'

    def __init__(self, file, class_no):
        super().__init__(file, class_no)

    def core_model(self):
        predicted = []
        error = []
        for i in range(len(self.X_test)):
            # Rebuild Model every iteration
            model = ARIMA(self.X_train, order = (5,1,0)).fit(disp=0)

            # Predict test data
            predicted.append(model.forecast()[0])
            
            # Update training data
            self.X_train.append(self.X_test[i])
            
            #Calculate error
            error.append(math.fabs(predicted[i][0] - self.X_test[i]))

            #print('%f, predicted=%f, expected=%f, error=%f' % (i, predicted[i][0], self.X_test[i], error[i]))

        self.mse = self.get_mse(self.X_test, predicted)
        self.mae = self.get_mae(self.X_test, predicted)
        
        #To save model
        model.save('arima_model.pkl')

        print("Mean Squared error = {:.4f}".format(self.mse))
        print("Mean Absolute error = {:.4f}".format(self.mae))

class HoltWinter_Model(Model):

    name = 'Holtzmann-Winter'

    def __init__(self, file, class_no):
        self.alpha, self.beta, self.gamma = 0.1, 0.1, 0.1
        self.season_length = 4
        super().__init__(file, class_no)

    def initial_trend(self, X_train, season_length):
        total = 0
        for i in range(season_length):
            t = (X_train[i+season_length] - X_train[i])
            t /= season_length
            total += t
        return total / season_length

    def initial_season_components(self, X_train, season_length):
        no_of_seasons = int(len(X_train)/season_length)
        season_averages = []
        seasonal_data = {}
        
        # Find season averages
        for season in range(no_of_seasons):
            i = season_length*season
            season_data = X_train[i:i+season_length]
            season_avg = sum(season_data)/season_length
            season_averages.append(season_avg)
            
        # Find initial values for each season
        for i in range(season_length):
            total_over_avg = 0
            for season in range(no_of_seasons):
                ind = season_length*season
                total_over_avg += X_train[ind+i] - season_averages[season]
            seasonal_data[i] = total_over_avg/no_of_seasons
        return seasonal_data

    def triple_exponential_smoothing(self, X_train, X_test, season_length, alpha, beta, gamma, no_to_predict):
        seasonal_data = self.initial_season_components(X_train, season_length)
            
        smooth = X_train[0]
        trend = self.initial_trend(X_train, season_length)
        result = [X_train[0]]
        
        for i in range(season_length + no_to_predict):
            if i >= len(X_train):
                m = i - len(X_train) + 1
                val = smooth + m * trend + seasonal_data[i % season_length]
                result.append(val)
            else:
                val = X_train[i]
                last_smooth, smooth = smooth, alpha * (val - seasonal_data [i%season_length]) + (1-alpha) * (smooth + trend)
                trend = beta * (smooth - last_smooth) + (1-beta) * trend
                seasonal_data[i % season_length] = gamma * (val-smooth) + (1-gamma) * seasonal_data[i%season_length]
                tot = smooth + trend + seasonal_data[i%season_length]
                result.append(tot)
        return result

    def core_model(self):
        no_to_predict = len(self.X_train) + len(self.X_test)
        result = self.triple_exponential_smoothing(self.X_train, self.X_test, self.season_length, self.alpha, self.beta, self.gamma, no_to_predict)

        self.mse = self.get_mse(self.X_train, result[:-len(self.X_test) - self.season_length - 1])
        self.mae = self.get_mae(self.X_train, result[:-len(self.X_test) - self.season_length - 1])

        # self.mse = self.get_mse(self.X_test, result[:-self.season_length - 1])
        # self.mae = self.get_mae(self.X_test, result[:-self.season_length - 1])
        
        print("Mean Squared error = {:.4f}".format(self.mse))
        print("Mean Absolute error = {:.4f}".format(self.mae))
    

#Starter
class_no = sys.argv[1]

file = 'SalesData.csv'
model_runs = [ARIMA_Model(file, class_no), HoltWinter_Model(file, class_no)]
min_mse = model_runs[0].mse
best_model = model_runs[0].name
for model in model_runs:
    if min_mse > model.mse:
        min_mse = model.mse
        best_model = model.name

print(best_model)